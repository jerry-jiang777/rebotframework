# !/usr/bin python3
# encoding: utf-8 -*-
# @file     : file_load.py
# @author   : 沙陌 Matongxue_2
# @Time     : 2023/7/22 14:42
# @Copyright: 北京码同学
import openpyxl
import yaml

from paths_manager import mtxshop_data_yaml


def read_excel(filepath,sheet_name):
    # 获取整个文档对象
    wb = openpyxl.load_workbook(filepath)
    # 获取某个sheet工作表的数据
    sheet_data = wb[sheet_name]
    # print(sheet_data)
    # 获取某个单元格数据
    # print(sheet_data.cell(2, 2).value)
    lines_count = sheet_data.max_row # 获取总行数
    cols_count = sheet_data.max_column # 获取总列数
    # print(lines_count,cols_count)
    # 注意：openpyxl里读取时行号和列号都从是1开始
    data = [] # 用来存储所有行的数据，每行数据都是这个列表的子列表
    for l in range(2,lines_count+1): # l:2,3,4,5
        line = [] # 用来存储当前行所有列的单元格数据
        for c in range(1,cols_count+1): # c:1,2,3,4,5
            # print(sheet_data.cell(l,c).value)
            cell_data = sheet_data.cell(l,c).value
            line.append(cell_data)
        data.append(line)
    return data

def load_yaml_file(filepath):
    with open(file=filepath,mode='r',encoding='utf-8') as f:
        content = yaml.load(f,Loader=yaml.FullLoader)
        return content

def write_yaml(filepath: object, content: object) -> None:
    with open(file=filepath, mode='w', encoding='utf-8') as f:
        yaml.dump(content, f, Dumper=yaml.Dumper)

if __name__ == '__main__':
    print(read_excel('../data/mtxshop_data.xlsx', '立即购买测试数据'))
    print(load_yaml_file(mtxshop_data_yaml))



